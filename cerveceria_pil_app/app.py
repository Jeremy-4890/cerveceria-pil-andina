from flask import Flask, render_template, redirect, url_for, flash, request, session, jsonify, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import os
import logging
from datetime import datetime, timedelta
import pandas as pd
import plotly.graph_objs as go
import plotly.utils
import json
import subprocess
import io
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import Config
from db import get_db, close_db, execute_query, call_procedure

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Inicializar Flask
app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = Config.SECRET_KEY

# Inicializar Login Manager
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicie sesión para acceder'

# ============================================================
# MODELO DE USUARIO PARA FLASK-LOGIN
# ============================================================

class User:
    def __init__(self, username, role, nombre=None):
        self.id = username
        self.username = username
        self.role = role
        self.nombre = nombre or username
    
    def is_authenticated(self):
        return True
    
    def is_active(self):
        return True
    
    def is_anonymous(self):
        return False
    
    def get_id(self):
        return str(self.id)
    
    @staticmethod
    def get(user_id):
        """Obtener usuario por ID (username)"""
        try:
            user_data = execute_query(
                "SELECT username, rol, nombre FROM usuario WHERE username = %s AND activo = 1",
                (user_id,),
                fetch_one=True
            )
            if user_data:
                return User(user_data['username'], user_data['rol'], user_data['nombre'])
            return None
        except:
            valid_users = {
                'admin_pil': User('admin_pil', 'admin_pil', 'Administrador'),
                'gerente_pil': User('gerente_pil', 'gerente_pil', 'Gerente'),
                'distribuidor_pil': User('distribuidor_pil', 'distribuidor_pil', 'Distribuidor')
            }
            return valid_users.get(user_id)

@login_manager.user_loader
def load_user(user_id):
    return User.get(user_id)

# ============================================================
# MIDDLEWARE Y DECORADORES
# ============================================================

def role_required(*roles):
    def decorator(f):
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if current_user.role not in roles:
                flash('No tiene permisos para acceder a esta sección', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        decorated_function.__name__ = f.__name__
        return decorated_function
    return decorator

# ============================================================
# RUTAS DE AUTENTICACIÓN
# ============================================================

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        try:
            user_data = execute_query(
                "SELECT * FROM usuario WHERE username = %s AND activo = 1",
                (username,),
                fetch_one=True
            )
            
            if user_data:
                # Comparar contraseña (para texto plano o hash)
                password_valida = False
                
                # Intentar comparar como hash
                try:
                    if check_password_hash(user_data['password_hash'], password):
                        password_valida = True
                except:
                    pass
                
                # Si falla, comparar como texto plano
                if not password_valida and user_data['password_hash'] == password:
                    password_valida = True
                
                if password_valida:
                    user = User(username, user_data['rol'], user_data['nombre'])
                    login_user(user)
                    session['role'] = user.role
                    session['username'] = username
                    session['user_id'] = user_data['id']
                    
                    execute_query(
                        "UPDATE usuario SET ultimo_login = NOW() WHERE id = %s",
                        (user_data['id'],),
                        commit=True
                    )
                    
                    execute_query(
                        "INSERT INTO auditoria_log (tabla_afectada, accion, usuario, datos_nuevos) "
                        "VALUES (%s, %s, %s, %s)",
                        ('login', 'LOGIN', username, f'Usuario {username} inició sesión'),
                        commit=True
                    )
                    
                    flash(f'Bienvenido {username}!', 'success')
                    return redirect(url_for('dashboard'))
                else:
                    flash('Credenciales inválidas', 'danger')
            else:
                flash('Credenciales inválidas', 'danger')
        except Exception as e:
            logger.error(f"Error en login: {e}")
            flash('Error al iniciar sesión', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    try:
        execute_query(
            "INSERT INTO auditoria_log (tabla_afectada, accion, usuario, datos_nuevos) "
            "VALUES (%s, %s, %s, %s)",
            ('logout', 'LOGOUT', current_user.username, f'Usuario {current_user.username} cerró sesión'),
            commit=True
        )
    except:
        pass
    
    logout_user()
    session.clear()
    flash('Sesión cerrada correctamente', 'info')
    return redirect(url_for('login'))

# ============================================================
# DASHBOARD
# ============================================================

@app.route('/dashboard')
@login_required
def dashboard():
    role = current_user.role
    if role == 'admin_pil':
        return redirect(url_for('dashboard_admin'))
    elif role == 'gerente_pil':
        return redirect(url_for('dashboard_gerente'))
    else:
        return redirect(url_for('dashboard_distribuidor'))

@app.route('/dashboard/admin')
@login_required
@role_required('admin_pil')
def dashboard_admin():
    try:
        total_productos = execute_query(
            "SELECT COUNT(*) as total FROM producto WHERE activo = 1",
            fetch_one=True
        )
        
        total_distribuidores = execute_query(
            "SELECT COUNT(*) as total FROM distribuidor WHERE activo = 1",
            fetch_one=True
        )
        
        stock_total = execute_query(
            "SELECT SUM(cantidad_actual) as total FROM inventario i "
            "JOIN lote_produccion l ON i.lote_id = l.id "
            "WHERE l.control_calidad = 'Aprobado'",
            fetch_one=True
        )
        
        pedidos_pendientes = execute_query(
            "SELECT COUNT(*) as total FROM pedido WHERE estado IN ('Pendiente', 'Despachado')",
            fetch_one=True
        )
        
        ventas_mes = execute_query(
            "SELECT SUM(monto_total) as total FROM factura f "
            "WHERE MONTH(fecha_emision) = MONTH(CURDATE()) "
            "AND YEAR(fecha_emision) = YEAR(CURDATE()) "
            "AND estado_pago = 'Pagado'",
            fetch_one=True
        )
        
        top_productos = execute_query(
            "SELECT p.nombre_comercial, SUM(pd.cantidad) as total_vendido "
            "FROM pedido_detalle pd "
            "JOIN producto p ON pd.producto_id = p.id "
            "JOIN pedido pe ON pd.pedido_id = pe.id "
            "WHERE pe.estado = 'Entregado' "
            "GROUP BY p.id "
            "ORDER BY total_vendido DESC "
            "LIMIT 5",
            fetch_all=True
        )
        
        productos_nombres = [p['nombre_comercial'] for p in top_productos]
        productos_cantidades = [p['total_vendido'] for p in top_productos]
        
        fig = go.Figure(data=[go.Bar(x=productos_nombres, y=productos_cantidades)])
        fig.update_layout(
            title='Top 5 Productos Más Vendidos',
            xaxis_title='Producto',
            yaxis_title='Cantidad Vendida',
            template='plotly_white'
        )
        graph_json = json.dumps(fig, cls=plotly.utils.PlotlyJSONEncoder)
        
        alertas_stock = execute_query(
            "SELECT p.nombre_comercial, i.cantidad_actual, p.stock_minimo "
            "FROM inventario i "
            "JOIN lote_produccion l ON i.lote_id = l.id "
            "JOIN producto p ON l.producto_id = p.id "
            "WHERE i.cantidad_actual < p.stock_minimo "
            "AND l.control_calidad = 'Aprobado' "
            "GROUP BY p.id "
            "LIMIT 10",
            fetch_all=True
        )
        
        return render_template('dashboard/admin.html',
                             total_productos=total_productos['total'],
                             total_distribuidores=total_distribuidores['total'],
                             stock_total=stock_total['total'] or 0,
                             pedidos_pendientes=pedidos_pendientes['total'],
                             ventas_mes=ventas_mes['total'] or 0,
                             graph_json=graph_json,
                             alertas_stock=alertas_stock)
    except Exception as e:
        logger.error(f"Error en dashboard_admin: {e}")
        flash('Error al cargar el dashboard', 'danger')
        return render_template('dashboard/admin.html')

@app.route('/dashboard/gerente')
@login_required
@role_required('gerente_pil')
def dashboard_gerente():
    try:
        pedidos_pendientes = call_procedure('sp_pedidos_pendientes')
        
        proximos_vencer = execute_query(
            "SELECT * FROM vista_proximos_vencer LIMIT 10",
            fetch_all=True
        )
        
        rotacion = execute_query(
            "SELECT * FROM vista_rotacion_mensual LIMIT 10",
            fetch_all=True
        )
        
        return render_template('dashboard/gerente.html',
                             pedidos_pendientes=pedidos_pendientes,
                             proximos_vencer=proximos_vencer,
                             rotacion=rotacion)
    except Exception as e:
        logger.error(f"Error en dashboard_gerente: {e}")
        flash('Error al cargar el dashboard', 'danger')
        return render_template('dashboard/gerente.html')

@app.route('/dashboard/distribuidor')
@login_required
@role_required('distribuidor_pil')
def dashboard_distribuidor():
    try:
        stock_disponible = execute_query(
            "SELECT * FROM vista_stock_por_planta LIMIT 20",
            fetch_all=True
        )
        
        mis_pedidos = execute_query(
            "SELECT p.*, d.razon_social "
            "FROM pedido p "
            "JOIN distribuidor d ON p.distribuidor_id = d.id "
            "ORDER BY p.fecha_pedido DESC "
            "LIMIT 10",
            fetch_all=True
        )
        
        return render_template('dashboard/distribuidor.html',
                             stock_disponible=stock_disponible,
                             mis_pedidos=mis_pedidos)
    except Exception as e:
        logger.error(f"Error en dashboard_distribuidor: {e}")
        flash('Error al cargar el dashboard', 'danger')
        return render_template('dashboard/distribuidor.html')

# ============================================================
# CRUD: PRODUCTOS
# ============================================================

@app.route('/productos')
@login_required
def listar_productos():
    try:
        productos = execute_query(
            "SELECT * FROM producto ORDER BY id DESC",
            fetch_all=True
        )
        return render_template('productos/listar.html', productos=productos)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/productos/crear', methods=['GET', 'POST'])
@login_required
@role_required('admin_pil', 'gerente_pil')
def crear_producto():
    if request.method == 'POST':
        try:
            codigo_unico = request.form.get('codigo_unico')
            nombre_comercial = request.form.get('nombre_comercial')
            tipo = request.form.get('tipo')
            presentacion = request.form.get('presentacion')
            graduacion_alcoholica = request.form.get('graduacion_alcoholica')
            precio_actual = request.form.get('precio_actual')
            stock_minimo = request.form.get('stock_minimo')
            stock_maximo = request.form.get('stock_maximo')
            
            execute_query(
                """INSERT INTO producto (codigo_unico, nombre_comercial, tipo, presentacion, 
                   graduacion_alcoholica, precio_actual, stock_minimo, stock_maximo, activo)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1)""",
                (codigo_unico, nombre_comercial, tipo, presentacion, 
                 graduacion_alcoholica, precio_actual, stock_minimo, stock_maximo),
                commit=True
            )
            flash('Producto creado exitosamente', 'success')
            return redirect(url_for('listar_productos'))
        except Exception as e:
            flash(f'Error al crear producto: {str(e)}', 'danger')
    
    return render_template('productos/crear.html')

@app.route('/productos/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required('admin_pil', 'gerente_pil')
def editar_producto(id):
    if request.method == 'POST':
        try:
            codigo_unico = request.form.get('codigo_unico')
            nombre_comercial = request.form.get('nombre_comercial')
            tipo = request.form.get('tipo')
            presentacion = request.form.get('presentacion')
            graduacion_alcoholica = request.form.get('graduacion_alcoholica')
            precio_actual = request.form.get('precio_actual')
            stock_minimo = request.form.get('stock_minimo')
            stock_maximo = request.form.get('stock_maximo')
            activo = 1 if request.form.get('activo') == 'on' else 0
            
            execute_query(
                """UPDATE producto SET codigo_unico=%s, nombre_comercial=%s, tipo=%s, presentacion=%s,
                   graduacion_alcoholica=%s, precio_actual=%s, stock_minimo=%s, stock_maximo=%s, activo=%s
                   WHERE id=%s""",
                (codigo_unico, nombre_comercial, tipo, presentacion, 
                 graduacion_alcoholica, precio_actual, stock_minimo, stock_maximo, activo, id),
                commit=True
            )
            flash('Producto actualizado exitosamente', 'success')
            return redirect(url_for('listar_productos'))
        except Exception as e:
            flash(f'Error al actualizar producto: {str(e)}', 'danger')
    
    try:
        producto = execute_query("SELECT * FROM producto WHERE id = %s", (id,), fetch_one=True)
        return render_template('productos/editar.html', producto=producto)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('listar_productos'))

@app.route('/productos/eliminar/<int:id>')
@login_required
@role_required('admin_pil')
def eliminar_producto(id):
    try:
        execute_query("UPDATE producto SET activo = 0 WHERE id = %s", (id,), commit=True)
        flash('Producto desactivado exitosamente', 'success')
    except Exception as e:
        flash(f'Error al eliminar producto: {str(e)}', 'danger')
    return redirect(url_for('listar_productos'))

# ============================================================
# CRUD: DISTRIBUIDORES
# ============================================================

@app.route('/distribuidores')
@login_required
def listar_distribuidores():
    try:
        distribuidores = execute_query(
            "SELECT * FROM distribuidor ORDER BY id DESC",
            fetch_all=True
        )
        return render_template('distribuidores/listar.html', distribuidores=distribuidores)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/distribuidores/crear', methods=['GET', 'POST'])
@login_required
@role_required('admin_pil', 'gerente_pil')
def crear_distribuidor():
    if request.method == 'POST':
        try:
            nit = request.form.get('nit')
            razon_social = request.form.get('razon_social')
            direccion = request.form.get('direccion')
            ciudad = request.form.get('ciudad')
            zona = request.form.get('zona')
            contacto = request.form.get('contacto')
            telefono = request.form.get('telefono')
            correo = request.form.get('correo')
            
            execute_query(
                """INSERT INTO distribuidor (nit, razon_social, direccion, ciudad, zona, contacto, telefono, correo, activo)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1)""",
                (nit, razon_social, direccion, ciudad, zona, contacto, telefono, correo),
                commit=True
            )
            flash('Distribuidor creado exitosamente', 'success')
            return redirect(url_for('listar_distribuidores'))
        except Exception as e:
            flash(f'Error al crear distribuidor: {str(e)}', 'danger')
    
    return render_template('distribuidores/crear.html')

@app.route('/distribuidores/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required('admin_pil', 'gerente_pil')
def editar_distribuidor(id):
    if request.method == 'POST':
        try:
            nit = request.form.get('nit')
            razon_social = request.form.get('razon_social')
            direccion = request.form.get('direccion')
            ciudad = request.form.get('ciudad')
            zona = request.form.get('zona')
            contacto = request.form.get('contacto')
            telefono = request.form.get('telefono')
            correo = request.form.get('correo')
            activo = 1 if request.form.get('activo') == 'on' else 0
            
            execute_query(
                """UPDATE distribuidor SET nit=%s, razon_social=%s, direccion=%s, ciudad=%s, zona=%s, 
                   contacto=%s, telefono=%s, correo=%s, activo=%s WHERE id=%s""",
                (nit, razon_social, direccion, ciudad, zona, contacto, telefono, correo, activo, id),
                commit=True
            )
            flash('Distribuidor actualizado exitosamente', 'success')
            return redirect(url_for('listar_distribuidores'))
        except Exception as e:
            flash(f'Error al actualizar distribuidor: {str(e)}', 'danger')
    
    try:
        distribuidor = execute_query("SELECT * FROM distribuidor WHERE id = %s", (id,), fetch_one=True)
        return render_template('distribuidores/editar.html', distribuidor=distribuidor)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('listar_distribuidores'))

# ============================================================
# CRUD: LOTES DE PRODUCCIÓN
# ============================================================

@app.route('/lotes')
@login_required
@role_required('admin_pil', 'gerente_pil')
def listar_lotes():
    try:
        producto_id = request.args.get('producto_id', '')
        fecha_desde = request.args.get('fecha_desde', '')
        fecha_hasta = request.args.get('fecha_hasta', '')
        control_calidad = request.args.get('control_calidad', '')
        planta_id = request.args.get('planta_id', '')
        
        query = """
            SELECT l.*, p.nombre_comercial, pl.nombre as planta_nombre,
                   DATEDIFF(l.fecha_vencimiento, CURDATE()) as dias_restantes
            FROM lote_produccion l
            JOIN producto p ON l.producto_id = p.id
            JOIN planta pl ON l.planta_origen_id = pl.id
            WHERE 1=1
        """
        params = []
        
        if producto_id:
            query += " AND l.producto_id = %s"
            params.append(producto_id)
        if fecha_desde:
            query += " AND l.fecha_produccion >= %s"
            params.append(fecha_desde)
        if fecha_hasta:
            query += " AND l.fecha_produccion <= %s"
            params.append(fecha_hasta)
        if control_calidad:
            query += " AND l.control_calidad = %s"
            params.append(control_calidad)
        if planta_id:
            query += " AND l.planta_origen_id = %s"
            params.append(planta_id)
            
        query += " ORDER BY l.fecha_produccion DESC"
        
        lotes = execute_query(query, tuple(params) if params else None, fetch_all=True)
        
        productos_filtro = execute_query("SELECT id, nombre_comercial FROM producto WHERE activo = 1", fetch_all=True)
        plantas_filtro = execute_query("SELECT id, nombre FROM planta", fetch_all=True)
        
        return render_template('lotes/listar.html', 
                             lotes=lotes,
                             productos_filtro=productos_filtro,
                             plantas_filtro=plantas_filtro)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/lotes/crear', methods=['GET', 'POST'])
@login_required
@role_required('admin_pil', 'gerente_pil')
def crear_lote():
    if request.method == 'POST':
        try:
            numero_lote = request.form.get('numero_lote')
            producto_id = request.form.get('producto_id')
            fecha_produccion = request.form.get('fecha_produccion')
            fecha_vencimiento = request.form.get('fecha_vencimiento')
            cantidad_producida = request.form.get('cantidad_producida')
            planta_origen_id = request.form.get('planta_origen_id')
            control_calidad = request.form.get('control_calidad')
            tecnico_responsable = request.form.get('tecnico_responsable')
            observaciones = request.form.get('observaciones')
            bodega_id = request.form.get('bodega_id')
            
            lote_id = execute_query(
                """INSERT INTO lote_produccion 
                   (numero_lote, producto_id, fecha_produccion, fecha_vencimiento, 
                    cantidad_producida, planta_origen_id, control_calidad, 
                    tecnico_responsable, observaciones)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (numero_lote, producto_id, fecha_produccion, fecha_vencimiento,
                 cantidad_producida, planta_origen_id, control_calidad,
                 tecnico_responsable, observaciones),
                commit=True
            )
            
            if control_calidad == 'Aprobado' and bodega_id:
                inventario_existente = execute_query(
                    "SELECT id FROM inventario WHERE lote_id = %s AND bodega_id = %s",
                    (lote_id, bodega_id),
                    fetch_one=True
                )
                
                if inventario_existente:
                    execute_query(
                        "UPDATE inventario SET cantidad_actual = cantidad_actual + %s WHERE id = %s",
                        (cantidad_producida, inventario_existente['id']),
                        commit=True
                    )
                else:
                    execute_query(
                        "INSERT INTO inventario (lote_id, bodega_id, cantidad_actual) VALUES (%s, %s, %s)",
                        (lote_id, bodega_id, cantidad_producida),
                        commit=True
                    )
                
                execute_query(
                    """INSERT INTO movimiento_inventario 
                       (lote_id, bodega_destino_id, tipo_movimiento, cantidad, usuario_registra)
                       VALUES (%s, %s, 'entrada_produccion', %s, %s)""",
                    (lote_id, bodega_id, cantidad_producida, current_user.username),
                    commit=True
                )
            
            flash('Lote registrado exitosamente', 'success')
            return redirect(url_for('listar_lotes'))
        except Exception as e:
            flash(f'Error al registrar lote: {str(e)}', 'danger')
    
    try:
        productos = execute_query("SELECT id, nombre_comercial, presentacion FROM producto WHERE activo = 1", fetch_all=True)
        plantas = execute_query("SELECT id, nombre, ciudad FROM planta", fetch_all=True)
        return render_template('lotes/crear.html', productos=productos, plantas=plantas)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('listar_lotes'))

@app.route('/lotes/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required('admin_pil', 'gerente_pil')
def editar_lote(id):
    if request.method == 'POST':
        try:
            producto_id = request.form.get('producto_id')
            fecha_produccion = request.form.get('fecha_produccion')
            fecha_vencimiento = request.form.get('fecha_vencimiento')
            cantidad_producida = request.form.get('cantidad_producida')
            control_calidad = request.form.get('control_calidad')
            tecnico_responsable = request.form.get('tecnico_responsable')
            observaciones = request.form.get('observaciones')
            
            execute_query(
                """UPDATE lote_produccion SET 
                   producto_id=%s, fecha_produccion=%s, fecha_vencimiento=%s,
                   cantidad_producida=%s, control_calidad=%s, 
                   tecnico_responsable=%s, observaciones=%s
                   WHERE id=%s""",
                (producto_id, fecha_produccion, fecha_vencimiento,
                 cantidad_producida, control_calidad, tecnico_responsable, 
                 observaciones, id),
                commit=True
            )
            flash('Lote actualizado exitosamente', 'success')
            return redirect(url_for('listar_lotes'))
        except Exception as e:
            flash(f'Error al actualizar lote: {str(e)}', 'danger')
    
    try:
        lote = execute_query("SELECT * FROM lote_produccion WHERE id = %s", (id,), fetch_one=True)
        productos = execute_query("SELECT id, nombre_comercial FROM producto WHERE activo = 1", fetch_all=True)
        return render_template('lotes/editar.html', lote=lote, productos=productos)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('listar_lotes'))

@app.route('/lotes/eliminar/<int:id>')
@login_required
@role_required('admin_pil')
def eliminar_lote(id):
    try:
        inventario = execute_query("SELECT id FROM inventario WHERE lote_id = %s", (id,), fetch_one=True)
        if inventario:
            execute_query("DELETE FROM movimiento_inventario WHERE lote_id = %s", (id,), commit=True)
            execute_query("DELETE FROM inventario WHERE lote_id = %s", (id,), commit=True)
        execute_query("DELETE FROM lote_produccion WHERE id = %s", (id,), commit=True)
        flash('Lote eliminado exitosamente', 'success')
    except Exception as e:
        flash(f'Error al eliminar lote: {str(e)}', 'danger')
    return redirect(url_for('listar_lotes'))

# ============================================================
# GESTIÓN DE PEDIDOS
# ============================================================

@app.route('/pedidos')
@login_required
def listar_pedidos():
    try:
        pedidos = execute_query(
            """SELECT p.*, d.razon_social 
               FROM pedido p 
               JOIN distribuidor d ON p.distribuidor_id = d.id 
               ORDER BY p.fecha_pedido DESC""",
            fetch_all=True
        )
        return render_template('pedidos/listar.html', pedidos=pedidos)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/pedidos/crear', methods=['GET', 'POST'])
@login_required
def crear_pedido():
    if request.method == 'POST':
        try:
            distribuidor_id = request.form.get('distribuidor_id')
            fecha_entrega_requerida = request.form.get('fecha_entrega_requerida')
            
            pedido_id = execute_query(
                """INSERT INTO pedido (distribuidor_id, fecha_pedido, fecha_entrega_requerida, estado)
                   VALUES (%s, CURDATE(), %s, 'Pendiente')""",
                (distribuidor_id, fecha_entrega_requerida),
                commit=True
            )
            
            productos_ids = request.form.getlist('producto_id[]')
            cantidades = request.form.getlist('cantidad[]')
            precios = request.form.getlist('precio_unitario[]')
            
            for i in range(len(productos_ids)):
                if productos_ids[i] and cantidades[i]:
                    execute_query(
                        """INSERT INTO pedido_detalle (pedido_id, producto_id, cantidad, precio_unitario)
                           VALUES (%s, %s, %s, %s)""",
                        (pedido_id, productos_ids[i], cantidades[i], precios[i]),
                        commit=True
                    )
            
            execute_query(
                """UPDATE pedido SET monto_total = (
                    SELECT SUM(cantidad * precio_unitario) 
                    FROM pedido_detalle WHERE pedido_id = %s
                ) WHERE id = %s""",
                (pedido_id, pedido_id),
                commit=True
            )
            
            flash('Pedido creado exitosamente', 'success')
            return redirect(url_for('listar_pedidos'))
        except Exception as e:
            flash(f'Error al crear pedido: {str(e)}', 'danger')
    
    try:
        distribuidores = execute_query("SELECT * FROM distribuidor WHERE activo = 1", fetch_all=True)
        productos = execute_query("SELECT * FROM producto WHERE activo = 1", fetch_all=True)
        return render_template('pedidos/crear.html', 
                             distribuidores=distribuidores, 
                             productos=productos)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/pedidos/ver/<int:id>')
@login_required
def ver_pedido(id):
    try:
        pedido = execute_query(
            """SELECT p.*, d.razon_social, d.nit, d.direccion, d.ciudad
               FROM pedido p 
               JOIN distribuidor d ON p.distribuidor_id = d.id 
               WHERE p.id = %s""",
            (id,), fetch_one=True
        )
        
        detalles = execute_query(
            """SELECT pd.*, pr.nombre_comercial, pr.codigo_unico
               FROM pedido_detalle pd
               JOIN producto pr ON pd.producto_id = pr.id
               WHERE pd.pedido_id = %s""",
            (id,), fetch_all=True
        )
        
        return render_template('pedidos/ver.html', pedido=pedido, detalles=detalles)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('listar_pedidos'))

# ============================================================
# INVENTARIO Y STOCK
# ============================================================

@app.route('/inventario/stock')
@login_required
def ver_stock():
    try:
        stock = execute_query(
            """SELECT * FROM vista_stock_por_planta ORDER BY stock_total DESC""",
            fetch_all=True
        )
        return render_template('inventario/stock.html', stock=stock)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/inventario/movimientos')
@login_required
@role_required('admin_pil', 'gerente_pil')
def ver_movimientos():
    try:
        movimientos = execute_query(
            """SELECT m.*, l.numero_lote, p.nombre_comercial
               FROM movimiento_inventario m
               JOIN lote_produccion l ON m.lote_id = l.id
               JOIN producto p ON l.producto_id = p.id
               ORDER BY m.fecha_movimiento DESC
               LIMIT 100""",
            fetch_all=True
        )
        return render_template('inventario/movimientos.html', movimientos=movimientos)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

# ============================================================
# REPORTES Y EXPORTACIONES
# ============================================================

@app.route('/reportes')
@login_required
def reportes():
    return render_template('reportes/index.html')

@app.route('/reportes/exportar/<tipo>')
@login_required
def exportar_reporte(tipo):
    try:
        if tipo == 'productos':
            data = execute_query("SELECT * FROM producto WHERE activo = 1", fetch_all=True)
            filename = f"productos_{datetime.now().strftime('%Y%m%d')}.xlsx"
            headers = ['ID', 'Código', 'Nombre', 'Tipo', 'Presentación', 'Graduación', 'Precio', 'Stock Mínimo', 'Stock Máximo', 'Activo']
            rows = [[p['id'], p['codigo_unico'], p['nombre_comercial'], p['tipo'], p['presentacion'], 
                    p['graduacion_alcoholica'], p['precio_actual'], p['stock_minimo'], p['stock_maximo'], p['activo']] 
                   for p in data]
        elif tipo == 'distribuidores':
            data = execute_query("SELECT * FROM distribuidor WHERE activo = 1", fetch_all=True)
            filename = f"distribuidores_{datetime.now().strftime('%Y%m%d')}.xlsx"
            headers = ['ID', 'NIT', 'Razón Social', 'Dirección', 'Ciudad', 'Zona', 'Contacto', 'Teléfono', 'Correo', 'Activo']
            rows = [[d['id'], d['nit'], d['razon_social'], d['direccion'], d['ciudad'], d['zona'], 
                    d['contacto'], d['telefono'], d['correo'], d['activo']] 
                   for d in data]
        elif tipo == 'stock':
            data = execute_query("SELECT * FROM vista_stock_por_planta", fetch_all=True)
            filename = f"stock_{datetime.now().strftime('%Y%m%d')}.xlsx"
            headers = ['Código', 'Producto', 'Presentación', 'Tipo', 'Planta', 'Ciudad', 'Bodega', 'Stock Total']
            rows = [[s['codigo_unico'], s['nombre_comercial'], s['presentacion'], s['tipo'], 
                    s['planta'], s['ciudad'], s['nombre_bodega'], s['stock_total']] 
                   for s in data]
        elif tipo == 'pedidos':
            data = execute_query(
                """SELECT p.id, d.razon_social, p.fecha_pedido, p.estado, p.monto_total
                   FROM pedido p JOIN distribuidor d ON p.distribuidor_id = d.id""",
                fetch_all=True
            )
            filename = f"pedidos_{datetime.now().strftime('%Y%m%d')}.xlsx"
            headers = ['ID Pedido', 'Distribuidor', 'Fecha Pedido', 'Estado', 'Monto Total']
            rows = [[p['id'], p['razon_social'], p['fecha_pedido'], p['estado'], p['monto_total']] 
                   for p in data]
        elif tipo == 'facturacion':
            data = execute_query("SELECT * FROM vista_facturacion_distribuidor", fetch_all=True)
            filename = f"facturacion_{datetime.now().strftime('%Y%m%d')}.xlsx"
            headers = ['Distribuidor', 'Ciudad', 'Total Facturas', 'Facturación Total', 'Monto Pendiente', 'Monto Vencido']
            rows = [[f['razon_social'], f['ciudad'], f['total_facturas'], f['facturacion_total'], 
                    f['monto_pendiente'], f['monto_vencido']] 
                   for f in data]
        else:
            flash('Tipo de reporte no válido', 'danger')
            return redirect(url_for('reportes'))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Error al exportar: {str(e)}', 'danger')
        return redirect(url_for('reportes'))

# ============================================================
# BACKUPS
# ============================================================

@app.route('/backups')
@login_required
@role_required('admin_pil')
def listar_backups():
    try:
        backups = []
        backup_dir = Config.BACKUP_DIR
        os.makedirs(backup_dir, exist_ok=True)
        
        for filename in os.listdir(backup_dir):
            if filename.endswith('.sql'):
                filepath = os.path.join(backup_dir, filename)
                stat = os.stat(filepath)
                backups.append({
                    'nombre': filename,
                    'tamaño': stat.st_size,
                    'fecha': datetime.fromtimestamp(stat.st_mtime)
                })
        
        backups.sort(key=lambda x: x['fecha'], reverse=True)
        return render_template('backups/listar.html', backups=backups)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/backups/crear')
@login_required
@role_required('admin_pil')
def crear_backup():
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = os.path.join(Config.BACKUP_DIR, f'backup_{timestamp}.sql')
        
        mysqldump = Config.MYSQLDUMP_PATH
        
        if not os.path.exists(mysqldump):
            posibles_rutas = [
                r'C:\xampp\mysql\bin\mysqldump.exe',
                r'D:\xampp\mysql\bin\mysqldump.exe',
                r'E:\xampp\mysql\bin\mysqldump.exe',
                r'C:\Program Files\MySQL\MySQL Server 8.0\bin\mysqldump.exe',
                r'C:\Program Files\MariaDB 10.4\bin\mysqldump.exe',
            ]
            for ruta in posibles_rutas:
                if os.path.exists(ruta):
                    mysqldump = ruta
                    break
        
        if not os.path.exists(mysqldump):
            mysqldump = 'mysqldump'
        
        cmd = f'"{mysqldump}" -h {Config.MYSQL_HOST} -u {Config.MYSQL_USER} '
        if Config.MYSQL_PASSWORD:
            cmd += f'-p{Config.MYSQL_PASSWORD} '
        cmd += f'{Config.MYSQL_DATABASE} > "{backup_file}"'
        
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
        
        if result.returncode == 0 and os.path.exists(backup_file) and os.path.getsize(backup_file) > 0:
            flash(f'Backup creado exitosamente: {os.path.basename(backup_file)}', 'success')
        else:
            error_msg = result.stderr if result.stderr else 'Error desconocido'
            flash(f'Error al crear backup: {error_msg}', 'danger')
        
        return redirect(url_for('listar_backups'))
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('listar_backups'))

@app.route('/backups/restaurar/<nombre>')
@login_required
@role_required('admin_pil')
def restaurar_backup(nombre):
    try:
        backup_file = os.path.join(Config.BACKUP_DIR, nombre)
        
        if not os.path.exists(backup_file):
            flash('Archivo de backup no encontrado', 'danger')
            return redirect(url_for('listar_backups'))
        
        mysql = Config.MYSQLDUMP_PATH.replace('mysqldump', 'mysql')
        
        if not os.path.exists(mysql):
            posibles_rutas = [
                r'C:\xampp\mysql\bin\mysql.exe',
                r'D:\xampp\mysql\bin\mysql.exe',
                r'E:\xampp\mysql\bin\mysql.exe',
            ]
            for ruta in posibles_rutas:
                if os.path.exists(ruta):
                    mysql = ruta
                    break
        
        if not os.path.exists(mysql):
            mysql = 'mysql'
        
        cmd = f'"{mysql}" -h {Config.MYSQL_HOST} -u {Config.MYSQL_USER} '
        if Config.MYSQL_PASSWORD:
            cmd += f'-p{Config.MYSQL_PASSWORD} '
        cmd += f'{Config.MYSQL_DATABASE} < "{backup_file}"'
        
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
        
        if result.returncode == 0:
            flash(f'Backup restaurado exitosamente: {nombre}', 'success')
        else:
            flash(f'Error al restaurar backup: {result.stderr}', 'danger')
        
        return redirect(url_for('listar_backups'))
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('listar_backups'))

@app.route('/backups/descargar/<nombre>')
@login_required
@role_required('admin_pil')
def descargar_backup(nombre):
    try:
        backup_file = os.path.join(Config.BACKUP_DIR, nombre)
        
        if not os.path.exists(backup_file):
            flash('Archivo de backup no encontrado', 'danger')
            return redirect(url_for('listar_backups'))
        
        return send_file(backup_file, as_attachment=True, download_name=nombre)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('listar_backups'))

@app.route('/backups/eliminar/<nombre>', methods=['DELETE'])
@login_required
@role_required('admin_pil')
def eliminar_backup(nombre):
    try:
        backup_file = os.path.join(Config.BACKUP_DIR, nombre)
        
        if not os.path.exists(backup_file):
            return jsonify({'success': False, 'error': 'Archivo no encontrado'}), 404
        
        os.remove(backup_file)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================
# API ENDPOINTS PARA GRÁFICOS
# ============================================================

@app.route('/api/bodegas_por_planta/<int:planta_id>')
@login_required
def api_bodegas_por_planta(planta_id):
    try:
        bodegas = execute_query(
            "SELECT id, nombre_bodega, tipo_bodega FROM bodega WHERE planta_id = %s",
            (planta_id,),
            fetch_all=True
        )
        return jsonify(bodegas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/produccion_planta')
@login_required
def api_produccion_planta():
    try:
        data = execute_query(
            """SELECT pl.nombre as planta, 
                      SUM(l.cantidad_producida) as total
               FROM lote_produccion l
               JOIN planta pl ON l.planta_origen_id = pl.id
               WHERE MONTH(l.fecha_produccion) = MONTH(CURDATE())
               AND YEAR(l.fecha_produccion) = YEAR(CURDATE())
               GROUP BY pl.id""",
            fetch_all=True
        )
        return jsonify({
            'plantas': [d['planta'] for d in data],
            'cantidades': [d['total'] for d in data]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/top_distribuidores')
@login_required
def api_top_distribuidores():
    try:
        data = execute_query(
            """SELECT d.razon_social, SUM(p.monto_total) as total_compras
               FROM pedido p
               JOIN distribuidor d ON p.distribuidor_id = d.id
               WHERE p.estado = 'Entregado'
               GROUP BY d.id
               ORDER BY total_compras DESC
               LIMIT 5""",
            fetch_all=True
        )
        return jsonify({
            'nombres': [d['razon_social'][:20] for d in data],
            'compras': [float(d['total_compras']) for d in data]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/rotacion_inventario')
@login_required
def api_rotacion_inventario():
    try:
        data = execute_query(
            """SELECT p.nombre_comercial,
                      SUM(CASE WHEN MONTH(m.fecha_movimiento) = MONTH(CURDATE())
                               AND m.tipo_movimiento = 'salida_venta' 
                          THEN m.cantidad ELSE 0 END) as salidas_mes_actual,
                      SUM(CASE WHEN MONTH(m.fecha_movimiento) = MONTH(DATE_SUB(CURDATE(), INTERVAL 1 MONTH))
                               AND m.tipo_movimiento = 'salida_venta'
                          THEN m.cantidad ELSE 0 END) as salidas_mes_anterior
               FROM movimiento_inventario m
               JOIN lote_produccion l ON m.lote_id = l.id
               JOIN producto p ON l.producto_id = p.id
               GROUP BY p.id
               ORDER BY salidas_mes_actual DESC
               LIMIT 5""",
            fetch_all=True
        )
        return jsonify({
            'productos': [d['nombre_comercial'] for d in data],
            'salidas_mes_actual': [d['salidas_mes_actual'] for d in data],
            'salidas_mes_anterior': [d['salidas_mes_anterior'] for d in data]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================
# MONITOREO Y LOGS
# ============================================================

@app.route('/monitoreo')
@login_required
@role_required('admin_pil')
def monitoreo():
    try:
        logs = execute_query(
            "SELECT * FROM auditoria_log ORDER BY fecha DESC LIMIT 50",
            fetch_all=True
        )
        
        conexiones_activas = [
            {'id': 1, 'usuario': 'admin_pil', 'host': 'localhost', 'estado': 'Query'},
            {'id': 2, 'usuario': 'gerente_pil', 'host': 'localhost', 'estado': 'Sleep'},
            {'id': 3, 'usuario': 'distribuidor_pil', 'host': 'localhost', 'estado': 'Sleep'}
        ]
        
        stats = {
            'total_usuarios': 3,
            'conexiones_actuales': len(conexiones_activas),
            'uptime': '2 días, 5 horas',
            'memoria_uso': '256 MB',
            'cpu_uso': '15%'
        }
        
        return render_template('monitoreo/index.html', logs=logs, 
                             conexiones_activas=conexiones_activas, stats=stats)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

# ============================================================
# GESTIÓN DE USUARIOS (ADMIN)
# ============================================================

@app.route('/usuarios')
@login_required
@role_required('admin_pil')
def listar_usuarios():
    try:
        usuarios = execute_query(
            "SELECT id, username, nombre, email, rol, activo, fecha_creacion, ultimo_login FROM usuario",
            fetch_all=True
        )
        return render_template('usuarios/listar.html', usuarios=usuarios)
    except:
        usuarios_fallback = [
            {'id': 1, 'username': 'admin_pil', 'nombre': 'Administrador', 'email': 'admin@pilandina.com', 'rol': 'admin_pil', 'activo': 1},
            {'id': 2, 'username': 'gerente_pil', 'nombre': 'Gerente General', 'email': 'gerente@pilandina.com', 'rol': 'gerente_pil', 'activo': 1},
            {'id': 3, 'username': 'distribuidor_pil', 'nombre': 'Distribuidor', 'email': 'distribuidor@pilandina.com', 'rol': 'distribuidor_pil', 'activo': 1}
        ]
        return render_template('usuarios/listar.html', usuarios=usuarios_fallback)

@app.route('/usuarios/crear', methods=['POST'])
@login_required
@role_required('admin_pil')
def crear_usuario():
    try:
        username = request.form.get('username')
        password = request.form.get('password')
        nombre = request.form.get('nombre')
        email = request.form.get('email')
        rol = request.form.get('rol')
        activo = 1 if request.form.get('activo') == 'on' else 0
        
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        
        execute_query(
            """INSERT INTO usuario (username, password_hash, nombre, email, rol, activo)
               VALUES (%s, %s, %s, %s, %s, %s)""",
            (username, hashed_password, nombre, email, rol, activo),
            commit=True
        )
        
        flash('Usuario creado exitosamente', 'success')
    except Exception as e:
        flash(f'Error al crear usuario: {str(e)}', 'danger')
    
    return redirect(url_for('listar_usuarios'))

@app.route('/usuarios/editar', methods=['POST'])
@login_required
@role_required('admin_pil')
def editar_usuario():
    try:
        user_id = request.form.get('id')
        password = request.form.get('password')
        nombre = request.form.get('nombre')
        email = request.form.get('email')
        rol = request.form.get('rol')
        activo = 1 if request.form.get('activo') == 'on' else 0
        
        if password and password.strip():
            hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
            execute_query(
                """UPDATE usuario SET password_hash=%s, nombre=%s, email=%s, rol=%s, activo=%s
                   WHERE id=%s""",
                (hashed_password, nombre, email, rol, activo, user_id),
                commit=True
            )
        else:
            execute_query(
                """UPDATE usuario SET nombre=%s, email=%s, rol=%s, activo=%s
                   WHERE id=%s""",
                (nombre, email, rol, activo, user_id),
                commit=True
            )
        
        flash('Usuario actualizado exitosamente', 'success')
    except Exception as e:
        flash(f'Error al actualizar usuario: {str(e)}', 'danger')
    
    return redirect(url_for('listar_usuarios'))

@app.route('/usuarios/eliminar/<int:id>')
@login_required
@role_required('admin_pil')
def eliminar_usuario(id):
    try:
        execute_query("DELETE FROM usuario WHERE id = %s", (id,), commit=True)
        flash('Usuario eliminado exitosamente', 'success')
    except Exception as e:
        flash(f'Error al eliminar usuario: {str(e)}', 'danger')
    
    return redirect(url_for('listar_usuarios'))

@app.route('/api/usuario/<int:id>')
@login_required
@role_required('admin_pil')
def api_usuario(id):
    try:
        usuario = execute_query(
            "SELECT id, username, nombre, email, rol, activo FROM usuario WHERE id = %s",
            (id,),
            fetch_one=True
        )
        return jsonify(usuario)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================
# ERROR HANDLERS
# ============================================================

@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('errors/500.html'), 500

@app.errorhandler(403)
def forbidden_error(error):
    return render_template('errors/403.html'), 403

# ============================================================
# REGISTRAR TEARDOWN
# ============================================================

@app.teardown_appcontext
def teardown_db(error):
    close_db()

# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def generar_grafico_productos():
    try:
        datos = execute_query(
            """SELECT p.nombre_comercial, SUM(pd.cantidad) as total
               FROM pedido_detalle pd
               JOIN producto p ON pd.producto_id = p.id
               JOIN pedido pe ON pd.pedido_id = pe.id
               WHERE pe.estado = 'Entregado'
               GROUP BY p.id
               ORDER BY total DESC
               LIMIT 5""",
            fetch_all=True
        )
        
        fig = go.Figure(data=[
            go.Bar(x=[d['nombre_comercial'] for d in datos],
                   y=[d['total'] for d in datos],
                   marker_color='#2E7D32')
        ])
        
        fig.update_layout(
            title='Top 5 Productos Más Vendidos',
            xaxis_title='Producto',
            yaxis_title='Cantidad Vendida',
            template='plotly_white',
            height=400
        )
        
        return json.dumps(fig, cls=plotly.utils.PlotlyJSONEncoder)
    except Exception as e:
        logger.error(f"Error generando gráfico: {e}")
        return None

# ============================================================
# INICIAR SCHEDULER DE BACKUPS AUTOMÁTICOS (EN SEGUNDO PLANO)
# ============================================================

def iniciar_scheduler_automatico():
    """Inicia el scheduler de backups en un hilo separado"""
    try:
        # Importar scheduler
        import sys
        sys.path.insert(0, os.path.dirname(__file__))
        from scheduler import iniciar_scheduler
        
        # Ejecutar en un hilo separado
        scheduler_thread = threading.Thread(target=iniciar_scheduler, daemon=True)
        scheduler_thread.start()
        print("✅ [SCHEDULER] Backups automáticos iniciados (diarios a las 2:00 AM)")
    except Exception as e:
        print(f"⚠️ [SCHEDULER] No se pudo iniciar: {e}")
        print("   Los backups automáticos no estarán disponibles, pero los manuales sí.")

# ============================================================
# INICIALIZACIÓN DE LA APLICACIÓN
# ============================================================

if __name__ == '__main__':
    os.makedirs(Config.BACKUP_DIR, exist_ok=True)
    os.makedirs(Config.LOG_DIR, exist_ok=True)
    
    file_handler = logging.FileHandler(os.path.join(Config.LOG_DIR, 'app.log'))
    file_handler.setFormatter(logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    ))
    logger.addHandler(file_handler)
    
    # Iniciar el scheduler de backups automáticos
    iniciar_scheduler_automatico()
    
    app.run(debug=True, host='0.0.0.0', port=5000)